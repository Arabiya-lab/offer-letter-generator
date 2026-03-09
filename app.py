"""
Reintenspark Technology — Offer Letter Generator (FINAL)
"""

import io, os, json, smtplib, re
from pathlib import Path
from zipfile import ZipFile
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication

import pandas as pd
import fitz
from flask import Flask, request, jsonify, send_file, abort
from reportlab.graphics.barcode.qr import QrCodeWidget
from reportlab.graphics.shapes import Drawing
from reportlab.graphics import renderPDF
from reportlab.pdfgen import canvas as rl_canvas
import openpyxl
from openpyxl import Workbook

app      = Flask(__name__)
BASE_DIR = Path(__file__).parent
DATA_DIR = BASE_DIR / "data"
DATA_DIR.mkdir(exist_ok=True)

EXCEL_FILE    = DATA_DIR / "interns.xlsx"
SETTINGS_FILE = DATA_DIR / "settings.json"

FONT_SIZE    = 12.0
LEFT_MARGIN  = 63.22
RIGHT_MARGIN = 542.0
LINE_HEIGHT  = 18.0

QR_SIZE     = 60
QR_RL_X     = 482.0
QR_RL_Y     = 704.0
QR_WIPE_X0  = 476.0
QR_WIPE_Y0  =  58.0
QR_WIPE_X1  = 558.0
QR_WIPE_Y1  = 145.0

REQUIRED_COLS = {"Name","Domain","Mode","Start_date","Duration","End_Date"}

# ── PUBLIC URL — set by ngrok at startup, used in every QR ────────────────
PUBLIC_URL = ""

import platform

def _find_font(reg_names, bold_names):
    dirs = []
    if platform.system() == "Windows":
        dirs = [r"C:\Windows\Fonts",
                os.path.join(os.environ.get("LOCALAPPDATA",""), "Microsoft","Windows","Fonts")]
    elif platform.system() == "Darwin":
        dirs = ["/Library/Fonts","/System/Library/Fonts", os.path.expanduser("~/Library/Fonts")]
    else:
        dirs = ["/usr/share/fonts","/usr/local/share/fonts", os.path.expanduser("~/.fonts")]
    for d in dirs:
        for rn in reg_names:
            rp = os.path.join(d, rn)
            if os.path.exists(rp):
                for bn in bold_names:
                    bp = os.path.join(d, bn)
                    if os.path.exists(bp):
                        return rp, bp
                return rp, None
    return None, None

_FONT_REG, _FONT_BOLD = _find_font(
    ["times.ttf","Times New Roman.ttf","TimesNewRoman.ttf","TIMES.TTF"],
    ["timesbd.ttf","Times New Roman Bold.ttf","TimesNewRomanBold.ttf","TIMESBD.TTF"]
)

def _font_kw(bold):
    if bold and _FONT_BOLD: return {"fontfile": _FONT_BOLD, "fontname": "tbd"}
    if not bold and _FONT_REG: return {"fontfile": _FONT_REG, "fontname": "treg"}
    return {"fontname": "Times-Bold" if bold else "Times-Roman"}

def _text_width(text, bold, fontsize):
    try:
        if bold and _FONT_BOLD: return fitz.Font(fontfile=_FONT_BOLD).text_length(text, fontsize=fontsize)
        if not bold and _FONT_REG: return fitz.Font(fontfile=_FONT_REG).text_length(text, fontsize=fontsize)
    except: pass
    try: return fitz.Font(fontname="Times-Bold" if bold else "Times-Roman").text_length(text, fontsize=fontsize)
    except: return len(text) * fontsize * 0.55

def init_excel():
    if not EXCEL_FILE.exists():
        wb = Workbook(); ws = wb.active; ws.title = "Interns"
        ws.append(["InternID","Name","Email","Date","Domain","Mode","Start_date","Duration","End_Date","Status"])
        wb.save(EXCEL_FILE)

def save_intern(d):
    init_excel()
    wb = openpyxl.load_workbook(EXCEL_FILE); ws = wb["Interns"]
    ws.append([d.get("InternID"), d.get("Name"), d.get("Email"), d.get("Date"),
               d.get("Domain"), d.get("Mode"), d.get("Start_date"),
               d.get("Duration"), d.get("End_Date"), d.get("Status","generated")])
    wb.save(EXCEL_FILE)

def fmt_date(v):
    try: return pd.to_datetime(v).strftime("%d %b %Y")
    except: return str(v)

def wipe(page, x0, y0, x1, y1):
    page.add_redact_annot(fitz.Rect(x0, y0, x1, y1))
    page.apply_redactions(images=fitz.PDF_REDACT_IMAGE_NONE)

def _draw_line(page, x, y, parts, fontsize):
    for text, bold in parts:
        if not text: continue
        kw = _font_kw(bold)
        page.insert_text((x, y), text, fontsize=fontsize, color=(0,0,0), overlay=True, **kw)
        x += _text_width(text, bold, fontsize)

def _draw_paragraph(page, x0, y, parts, fontsize, max_x, line_height):
    tokens = []
    for part_text, bold in parts:
        for piece in re.split(r"( +)", part_text):
            if piece: tokens.append((piece, bold))
    x = x0; at_line_start = True
    for token, bold in tokens:
        is_space = token.strip() == ""
        if at_line_start and is_space: continue
        w = _text_width(token, bold, fontsize)
        if not is_space and (x + w > max_x) and (x > x0):
            y += line_height; x = x0; at_line_start = True
            if is_space: continue
        kw = _font_kw(bold)
        page.insert_text((x, y), token, fontsize=fontsize, color=(0,0,0), overlay=True, **kw)
        x += w; at_line_start = False
    return y

def create_letter(intern, template_bytes, base_url="", include_qr=True):
    doc  = fitz.open(stream=template_bytes, filetype="pdf")
    page = doc.load_page(0)
    fs   = FONT_SIZE

    iid      = str(intern.get("InternID",""))
    name     = str(intern.get("Name",""))
    from datetime import date as _date
    date     = _date.today().strftime("%d/%m/%Y")
    domain   = str(intern.get("Domain",""))
    mode     = str(intern.get("Mode",""))
    start    = fmt_date(intern.get("Start_date",""))
    duration = str(intern.get("Duration",""))
    end      = fmt_date(intern.get("End_Date",""))

    wipe(page, QR_WIPE_X0, QR_WIPE_Y0, QR_WIPE_X1, QR_WIPE_Y1)
    wipe(page, 63.22, 193.0, 700.0, 212.0)
    _draw_line(page, 64.22, 208.73, [("Dear ", False),(name, True),(",", False)], fs)

    wipe(page, LEFT_MARGIN, 228.0, 700.0, 302.0)
    _draw_paragraph(page, LEFT_MARGIN, 244.14, [
        ("We are pleased to offer you an opportunity to join Reintenspark Technology Private Limited as an Intern with the Intern ID ", False),
        (iid, True), (" in the domain of ", False), (domain, True),
        (". Your internship will be conducted ", False), (mode, True),
        (" and will commence from ", False), (start, True),
        (" for a duration of ", False), (duration, True), (" month.", False),
    ], fs, RIGHT_MARGIN, LINE_HEIGHT)

    wipe(page, LEFT_MARGIN, 325.0, 700.0, 348.0)
    _draw_line(page, 79.51, 344.96, [("Mode: ", False),(mode, True),(".", False)], fs)

    wipe(page, LEFT_MARGIN, 348.0, 700.0, 367.0)
    _draw_line(page, 79.51, 362.96, [
        ("Duration: From ", False),(start, True),(" To ", False),(end, True),(".", False)], fs)

    wipe(page, 468.0, 137.0, 700.0, 161.0)
    _draw_line(page, 471.06, 157.07, [("Date: ", False),(date, True)], fs)

    tmp = io.BytesIO()
    doc.save(tmp, deflate=True, garbage=4, clean=True, pretty=False)
    doc.close(); tmp.seek(0)

    if include_qr:
        # ── QR encodes full verify URL so phone opens it as a clickable link ──
        # Priority: 1) base_url from UI  2) PUBLIC_URL from ngrok  3) LAN IP
        if base_url.strip():
            qr_data = base_url.rstrip("/") + "/verify/" + iid
        elif PUBLIC_URL.strip():
            qr_data = PUBLIC_URL.rstrip("/") + "/verify/" + iid
        else:
            import socket
            try:
                s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
                s.connect(("8.8.8.8", 80)); lan = s.getsockname()[0]; s.close()
            except: lan = "127.0.0.1"
            qr_data = f"http://{lan}:5000/verify/{iid}"

        fdoc_tmp = fitz.open(stream=tmp.getvalue(), filetype="pdf")
        rect = fdoc_tmp.load_page(0).rect; fdoc_tmp.close(); tmp.seek(0)

        ph = rect.height
        rl_y = ph - QR_WIPE_Y1 + (QR_WIPE_Y1 - QR_WIPE_Y0 - QR_SIZE) / 2

        pkt = io.BytesIO()
        c   = rl_canvas.Canvas(pkt, pagesize=(rect.width, rect.height))
        qw  = QrCodeWidget(qr_data)
        b   = qw.getBounds(); bw, bh = b[2]-b[0], b[3]-b[1]
        d   = Drawing(QR_SIZE, QR_SIZE, transform=[QR_SIZE/bw, 0, 0, QR_SIZE/bh, 0, 0])
        d.add(qw)
        renderPDF.draw(d, c, QR_RL_X, rl_y)
        c.save(); pkt.seek(0)

        letter_doc  = fitz.open(stream=tmp.getvalue(), filetype="pdf")
        qr_doc      = fitz.open(stream=pkt.getvalue(), filetype="pdf")
        letter_page = letter_doc.load_page(0)
        letter_page.show_pdf_page(letter_page.rect, qr_doc, 0, overlay=True)
        qr_doc.close()

        final_buf = io.BytesIO()
        letter_doc.save(final_buf, deflate=True, garbage=4, clean=True, pretty=False)
        letter_doc.close(); final_buf.seek(0)
        return final_buf.read()
    else:
        return tmp.read()

def send_email(cfg, recipient, intern_name, pdf_bytes, filename, subject, body):
    from sendgrid import SendGridAPIClient
from sendgrid.helpers.mail import Mail, Attachment, FileContent, FileName, FileType, Disposition
import base64
import os

def send_email(cfg, recipient, intern_name, pdf_bytes, filename, subject, body):

    body = body.replace("{{Name}}", intern_name)

    encoded_file = base64.b64encode(pdf_bytes).decode()

    attachment = Attachment(
        FileContent(encoded_file),
        FileName(filename),
        FileType("application/pdf"),
        Disposition("attachment")
    )

    message = Mail(
        from_email="hr@yourcompany.com",
        to_emails=recipient,
        subject=subject,
        plain_text_content=body
    )

    message.attachment = attachment

    try:
        sg = SendGridAPIClient(os.environ.get("SENDGRID_API_KEY"))
        sg.send(message)

    except Exception as e:
        raise Exception(f"SendGrid error: {e}")

_template_bytes = None

HTML = r"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Reintenspark — Offer Letter Generator</title>
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&family=JetBrains+Mono:wght@400;500&display=swap');
*{margin:0;padding:0;box-sizing:border-box}
:root{--bg:#f0f4f8;--surface:#ffffff;--border:#e2e8f0;--border2:#cbd5e1;--green:#5CB85C;--green-dark:#4cae4c;--green-light:#f0fdf4;--blue:#4f46e5;--blue-light:#eef2ff;--text:#1e293b;--text2:#475569;--text3:#94a3b8;--red:#ef4444;--yellow:#f59e0b;--mono:'JetBrains Mono',monospace;--shadow:0 1px 4px rgba(0,0,0,.08),0 4px 16px rgba(0,0,0,.04)}
body{font-family:'Inter',sans-serif;background:var(--bg);color:var(--text);min-height:100vh}
.header{background:linear-gradient(135deg,#0f172a 0%,#1e293b 100%);padding:24px 40px;display:flex;align-items:center;gap:18px;border-bottom:3px solid var(--green)}
.header-icon{font-size:40px}.header h1{color:#fff;font-size:22px;font-weight:700}.header h1 span{color:var(--green)}.header p{color:#94a3b8;font-size:13px;margin-top:3px}
.container{max-width:1100px;margin:0 auto;padding:32px 24px}
.grid2{display:grid;grid-template-columns:1fr 1fr;gap:20px}.grid3{display:grid;grid-template-columns:1fr 1fr 1fr;gap:16px}
.card{background:var(--surface);border:1px solid var(--border);border-radius:12px;padding:24px;box-shadow:var(--shadow);margin-bottom:20px}
.card-title{font-size:14px;font-weight:700;color:var(--text);margin-bottom:16px;display:flex;align-items:center;gap:8px}
.card-title .num{background:var(--green);color:#fff;width:22px;height:22px;border-radius:50%;display:flex;align-items:center;justify-content:center;font-size:11px;font-weight:800}
.field{display:flex;flex-direction:column;gap:6px;margin-bottom:14px}
label{font-size:11px;font-weight:700;color:var(--text3);text-transform:uppercase;letter-spacing:.7px}
input[type=text],input[type=email],input[type=number],input[type=password],select,textarea{background:#f8fafc;border:1.5px solid var(--border2);border-radius:8px;padding:9px 12px;font-family:inherit;font-size:13.5px;color:var(--text);outline:none;transition:border .15s,box-shadow .15s;width:100%}
input:focus,select:focus,textarea:focus{border-color:var(--green);box-shadow:0 0 0 3px rgba(92,184,92,.15)}
.upload-zone{border:2px dashed var(--border2);border-radius:10px;padding:28px;text-align:center;cursor:pointer;transition:all .2s;position:relative;background:#fafbfc}
.upload-zone:hover,.upload-zone.active{border-color:var(--green);background:var(--green-light)}
.upload-zone input{position:absolute;inset:0;opacity:0;cursor:pointer;width:100%;height:100%}
.upload-icon{font-size:32px;margin-bottom:8px}.upload-title{font-size:14px;font-weight:600;color:var(--text);margin-bottom:4px}.upload-sub{font-size:12px;color:var(--text3)}.upload-ok{font-size:12px;font-weight:600;color:var(--green);margin-top:8px}
.btn{display:inline-flex;align-items:center;gap:8px;padding:10px 20px;border-radius:8px;font-family:inherit;font-size:13.5px;font-weight:600;cursor:pointer;border:none;transition:all .15s;text-decoration:none}
.btn-primary{background:var(--green);color:#fff}.btn-primary:hover{background:var(--green-dark);transform:translateY(-1px)}.btn-primary:disabled{opacity:.5;cursor:not-allowed;transform:none}
.btn-secondary{background:var(--surface);color:var(--text2);border:1.5px solid var(--border2)}.btn-secondary:hover{border-color:var(--green);color:var(--green)}
.btn-full{width:100%;justify-content:center;padding:13px}.btn-sm{padding:6px 14px;font-size:12px}
.toggle-wrap{display:flex;align-items:center;gap:12px;padding:12px 0}
.toggle{position:relative;width:44px;height:24px}.toggle input{opacity:0;width:0;height:0}
.slider{position:absolute;inset:0;background:#cbd5e1;border-radius:24px;cursor:pointer;transition:.2s}
.slider:before{content:'';position:absolute;left:3px;top:3px;width:18px;height:18px;background:#fff;border-radius:50%;transition:.2s}
input:checked+.slider{background:var(--green)}input:checked+.slider:before{transform:translateX(20px)}
.toggle-label{font-size:13.5px;font-weight:500;color:var(--text)}
.pill{display:inline-block;padding:3px 10px;border-radius:20px;font-size:11px;font-weight:600;font-family:var(--mono)}
.pill-green{background:var(--green-light);color:#166534;border:1px solid #bbf7d0}.pill-blue{background:var(--blue-light);color:#3730a3;border:1px solid #c7d2fe}.pill-yellow{background:#fefce8;color:#854d0e;border:1px solid #fef08a}.pill-red{background:#fef2f2;color:#991b1b;border:1px solid #fecaca}
.alert{padding:12px 16px;border-radius:8px;font-size:13px;margin-bottom:14px;display:flex;align-items:flex-start;gap:10px}
.alert-info{background:var(--blue-light);color:#3730a3;border:1px solid #c7d2fe}.alert-success{background:var(--green-light);color:#166534;border:1px solid #bbf7d0}
.table-wrap{overflow-x:auto;border-radius:8px;border:1px solid var(--border)}
table{width:100%;border-collapse:collapse;font-size:13px}
thead th{background:#f8fafc;padding:9px 14px;text-align:left;font-size:11px;font-weight:700;color:var(--text3);text-transform:uppercase;letter-spacing:.7px;border-bottom:1px solid var(--border)}
tbody tr{border-bottom:1px solid var(--border)}tbody tr:last-child{border-bottom:none}tbody tr:hover{background:#f8fafc}
tbody td{padding:9px 14px;color:var(--text2);vertical-align:middle}td.name{color:var(--text);font-weight:500}
.progress-wrap{background:#e2e8f0;border-radius:4px;height:8px;overflow:hidden;margin:10px 0}
.progress-bar{height:100%;background:linear-gradient(90deg,var(--green),#86efac);border-radius:4px;transition:width .3s;width:0%}
.spinner{width:18px;height:18px;border:2px solid rgba(255,255,255,.3);border-top-color:#fff;border-radius:50%;animation:spin .6s linear infinite;display:inline-block}
@keyframes spin{to{transform:rotate(360deg)}}
.divider{height:1px;background:var(--border);margin:24px 0}
.result-item{display:flex;justify-content:space-between;align-items:center;padding:10px 14px;border-bottom:1px solid var(--border);font-size:13px}
.result-item:last-child{border-bottom:none}.result-name{font-weight:600;color:var(--text)}.result-id{font-family:var(--mono);font-size:11px;color:var(--text3)}
#toast-container{position:fixed;bottom:24px;right:24px;z-index:9999;display:flex;flex-direction:column;gap:8px}
.toast{background:#1e293b;color:#fff;padding:12px 18px;border-radius:10px;font-size:13px;box-shadow:0 8px 24px rgba(0,0,0,.3);animation:slideIn .2s ease;display:flex;align-items:center;gap:10px;max-width:340px}
.toast.success{border-left:3px solid var(--green)}.toast.error{border-left:3px solid var(--red)}.toast.info{border-left:3px solid var(--blue)}
@keyframes slideIn{from{transform:translateX(100%);opacity:0}to{transform:translateX(0);opacity:1}}
.col-tags{display:flex;flex-wrap:wrap;gap:6px;margin-top:10px}
.pub-url{background:#f0fdf4;border:1.5px solid #bbf7d0;border-radius:8px;padding:10px 14px;font-family:var(--mono);font-size:12px;color:#166534;margin-top:8px;word-break:break-all}
</style>
</head>
<body>
<div class="header">
  <div class="header-icon">📄</div>
  <div>
    <h1><span>REINTENSPARK</span> Offer Letter Generator</h1>
    <p>Upload intern Excel + PDF template → generate personalised letters with QR verification + auto email</p>
  </div>
</div>
<div class="container">
<div class="grid2">
  <div class="card">
    <div class="card-title"><span class="num">1</span> Upload Intern Excel</div>
    <div class="upload-zone" id="excelZone">
      <input type="file" id="excelFile" accept=".xlsx,.xls">
      <div class="upload-icon">📊</div>
      <div class="upload-title">Drop Excel file here or click to browse</div>
      <div class="upload-sub">.xlsx · .xls</div>
      <div class="upload-ok" id="excelOk" style="display:none"></div>
    </div>
    <div class="col-tags" style="margin-top:14px">
      <span style="font-size:11px;color:var(--text3);font-weight:600;align-self:center">Required columns:</span>
      <span class="pill pill-yellow">Name</span><span class="pill pill-yellow">Domain</span>
      <span class="pill pill-yellow">Mode</span><span class="pill pill-yellow">Start_date</span>
      <span class="pill pill-yellow">Duration</span><span class="pill pill-yellow">End_Date</span>
      <span class="pill pill-blue">Email</span>
      <span style="font-size:11px;color:var(--text3);align-self:center">(optional)</span>
    </div>
    <div style="margin-top:12px">
      <a href="/sample-excel" class="btn btn-secondary btn-sm">⬇ Download Sample Excel</a>
    </div>
  </div>
  <div class="card">
    <div class="card-title"><span class="num">2</span> Upload PDF Template</div>
    <div class="upload-zone" id="pdfZone">
      <input type="file" id="pdfFile" accept=".pdf">
      <div class="upload-icon">📄</div>
      <div class="upload-title">Drop your Reintenspark PDF template here</div>
      <div class="upload-sub">Your template design will be preserved exactly</div>
      <div class="upload-ok" id="pdfOk" style="display:none"></div>
    </div>
  </div>
</div>

<div class="card">
  <div class="card-title"><span class="num">3</span> Intern ID &amp; QR Settings</div>
  <div class="grid3">
    <div class="field"><label>ID Prefix</label><input type="text" id="idPrefix" value="RIS_"></div>
    <div class="field"><label>Starting Number</label><input type="number" id="idStart" value="1" min="1"></div>
    <div class="field"><label>Preview</label>
      <div style="padding:9px 12px;background:#f8fafc;border:1.5px solid var(--border2);border-radius:8px;font-family:var(--mono);font-size:13px;color:var(--blue)" id="idPreview">RIS_000001</div>
    </div>
  </div>
  <div class="field" style="margin-top:4px">
    <label>QR Base URL <span style="font-weight:400;text-transform:none">(leave blank to use ngrok auto-URL)</span></label>
    <input type="text" id="baseUrl" placeholder="Leave blank — ngrok URL is used automatically">
  </div>
  <div id="publicUrlBox" style="display:none">
    <div style="font-size:11px;font-weight:700;color:var(--text3);text-transform:uppercase;letter-spacing:.7px;margin-bottom:4px">Active QR URL (from ngrok)</div>
    <div class="pub-url" id="publicUrlText"></div>
  </div>
</div>

<div class="card">
  <div class="card-title"><span class="num">4</span> Email Settings <span style="font-size:12px;font-weight:400;color:var(--text3)">(optional)</span></div>
  <div class="toggle-wrap">
    <label class="toggle"><input type="checkbox" id="emailToggle" onchange="toggleEmail()"><span class="slider"></span></label>
    <span class="toggle-label">Automatically send offer letters to each intern via email</span>
  </div>
  <div id="emailSettings" style="display:none;margin-top:16px">
    <div class="alert alert-info">
      💡 For Gmail: use an <strong>App Password</strong> — <strong>myaccount.google.com → Security → App Passwords</strong>
    </div>
    <div class="grid2">
      <div class="field"><label>SMTP Host</label><input type="text" id="smtpHost" value="smtp.gmail.com"></div>
      <div class="field"><label>SMTP Port</label><input type="number" id="smtpPort" value="587"></div>
      <div class="field"><label>Sender Email</label><input type="email" id="smtpUser" placeholder="your@gmail.com"></div>
      <div class="field"><label>App Password</label><input type="password" id="smtpPass" placeholder="xxxx xxxx xxxx xxxx"></div>
    </div>
    <div class="field"><label>Email Subject</label>
      <input type="text" id="emailSubject" value="Your Internship Offer Letter – Reintenspark Technology Private Limited">
    </div>
    <div class="field">
      <label>Email Body <span style="font-weight:400;text-transform:none">(use {{Name}} as placeholder)</span></label>
      <textarea id="emailBody" rows="5">Dear {{Name}},

Please find attached your Internship Offer Letter from Reintenspark Technology Private Limited.

We look forward to having you on board!

Best regards,
Guru basha
HRM – Reintenspark Technology Private Limited
+91 8762719260</textarea>
    </div>
    <button class="btn btn-secondary" onclick="testEmail()" id="testEmailBtn">📧 Send Test Email</button>
    <div id="testEmailResult" style="margin-top:10px;font-size:13px"></div>
  </div>
</div>

<div class="card">
  <div class="card-title"><span class="num">5</span> Generate Offer Letters</div>
  <div id="previewSection" style="display:none;margin-bottom:16px">
    <div class="alert alert-success" id="previewAlert"></div>
    <div class="table-wrap" id="previewTable" style="margin-bottom:14px;max-height:300px;overflow-y:auto"></div>
  </div>
  <button class="btn btn-primary btn-full" id="generateBtn" onclick="generate()" disabled>⚡ Generate All Offer Letters</button>
  <div id="progressSection" style="display:none;margin-top:14px">
    <div style="font-size:13px;color:var(--text2);margin-bottom:6px" id="progressText">Processing…</div>
    <div class="progress-wrap"><div class="progress-bar" id="progressBar"></div></div>
  </div>
</div>

<div class="card" id="resultsCard" style="display:none">
  <div class="card-title"><span class="num">✓</span> Results</div>
  <div id="downloadAllWrap" style="margin-bottom:16px"></div>
  <div id="resultsList"></div>
  <div id="emailReport" style="margin-top:20px;display:none">
    <div class="divider"></div>
    <div class="card-title" style="margin-bottom:12px">📧 Email Delivery Report</div>
    <div class="grid3" id="emailStats"></div>
    <div class="table-wrap" style="margin-top:14px" id="emailTable"></div>
  </div>
</div>
</div>
<div id="toast-container"></div>
<script>
let excelData=null,pdfReady=false;
function toast(msg,type='info'){
  const icons={success:'✅',error:'❌',info:'ℹ️'};
  const el=document.createElement('div');el.className=`toast ${type}`;
  el.innerHTML=`<span>${icons[type]}</span> ${msg}`;
  document.getElementById('toast-container').appendChild(el);
  setTimeout(()=>el.remove(),5000);
}
// Load public URL from server and show it in UI
fetch('/api/public-url').then(r=>r.json()).then(d=>{
  if(d.url){
    document.getElementById('publicUrlBox').style.display='block';
    document.getElementById('publicUrlText').textContent=d.url+'/verify/RIS_XXXXXX';
  }
});
function updatePreview(){
  const pfx=document.getElementById('idPrefix').value;
  const num=parseInt(document.getElementById('idStart').value)||1;
  document.getElementById('idPreview').textContent=`${pfx}${String(num).padStart(6,'0')} · ${pfx}${String(num+1).padStart(6,'0')} …`;
}
document.getElementById('idPrefix').addEventListener('input',updatePreview);
document.getElementById('idStart').addEventListener('input',updatePreview);
function toggleEmail(){document.getElementById('emailSettings').style.display=document.getElementById('emailToggle').checked?'block':'none';}
function smtpPayload(){return{host:document.getElementById('smtpHost').value.trim(),port:document.getElementById('smtpPort').value.trim(),user:document.getElementById('smtpUser').value.trim(),pass:document.getElementById('smtpPass').value,subject:document.getElementById('emailSubject').value,body:document.getElementById('emailBody').value};}
async function testEmail(){
  const btn=document.getElementById('testEmailBtn'),res_el=document.getElementById('testEmailResult');
  btn.disabled=true;btn.innerHTML='<span class="spinner" style="border-top-color:var(--green);border-color:rgba(0,0,0,.15)"></span> Sending…';res_el.textContent='';
  try{
    const r=await fetch('/api/test-email',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({smtp:smtpPayload()})});
    const d=await r.json();
    if(d.success){res_el.innerHTML='<span style="color:var(--green);font-weight:600">✅ Test email sent!</span>';toast('Test email sent!','success');}
    else{res_el.innerHTML=`<span style="color:var(--red);font-weight:600">❌ ${d.error}</span>`;toast('Test failed: '+d.error,'error');}
  }catch(e){res_el.innerHTML=`<span style="color:var(--red)">❌ ${e.message}</span>`;}
  finally{btn.disabled=false;btn.innerHTML='📧 Send Test Email';}
}
function setupZone(zoneId,inputId,handler){
  const zone=document.getElementById(zoneId),input=document.getElementById(inputId);
  zone.addEventListener('dragover',e=>{e.preventDefault();zone.classList.add('active')});
  zone.addEventListener('dragleave',()=>zone.classList.remove('active'));
  zone.addEventListener('drop',e=>{e.preventDefault();zone.classList.remove('active');if(e.dataTransfer.files[0]){input.files=e.dataTransfer.files;handler(input.files[0])}});
  input.addEventListener('change',()=>{if(input.files[0])handler(input.files[0])});
}
setupZone('excelZone','excelFile',handleExcel);
setupZone('pdfZone','pdfFile',handlePdf);
async function handleExcel(file){
  const okEl=document.getElementById('excelOk');okEl.style.display='none';
  const fd=new FormData();fd.append('file',file);
  try{const res=await fetch('/api/parse-excel',{method:'POST',body:fd});const data=await res.json();
    if(data.error){toast(data.error,'error');return;}
    excelData=data.interns;okEl.textContent=`✅ ${data.count} intern(s) loaded`;okEl.style.display='block';
    showPreview(data.interns);checkReady();toast(`${data.count} intern(s) parsed`,'success');
  }catch(e){toast('Failed: '+e.message,'error');}
}
async function handlePdf(file){
  const okEl=document.getElementById('pdfOk');okEl.style.display='none';
  const fd=new FormData();fd.append('file',file);
  try{const res=await fetch('/api/upload-template',{method:'POST',body:fd});const data=await res.json();
    if(data.error){toast(data.error,'error');return;}
    pdfReady=true;okEl.textContent='✅ Template uploaded';okEl.style.display='block';
    checkReady();toast('PDF template ready','success');
  }catch(e){toast('Failed: '+e.message,'error');}
}
function checkReady(){document.getElementById('generateBtn').disabled=!(excelData&&pdfReady);}
function showPreview(interns){
  document.getElementById('previewSection').style.display='block';
  document.getElementById('previewAlert').textContent=`✅ ${interns.length} intern(s) ready`;
  const cols=['Name','Domain','Mode','Start_date','End_Date','Duration'];
  let html=`<table><thead><tr>${cols.map(c=>`<th>${c}</th>`).join('')}</tr></thead><tbody>`;
  interns.forEach(r=>{html+=`<tr>${cols.map(c=>`<td class="${c==='Name'?'name':''}">${r[c]||'—'}</td>`).join('')}</tr>`;});
  html+='</tbody></table>';
  document.getElementById('previewTable').innerHTML=html;
}
async function generate(){
  if(!excelData||!pdfReady)return;
  const btn=document.getElementById('generateBtn');
  btn.disabled=true;btn.innerHTML='<span class="spinner"></span> Generating…';
  const emailOn=document.getElementById('emailToggle').checked;
  const payload={interns:excelData,id_prefix:document.getElementById('idPrefix').value,id_start:parseInt(document.getElementById('idStart').value)||1,base_url:document.getElementById('baseUrl').value,send_email:emailOn,smtp:emailOn?smtpPayload():null};
  document.getElementById('progressSection').style.display='block';
  document.getElementById('progressBar').style.width='10%';
  document.getElementById('progressText').textContent='Generating…';
  try{
    const res=await fetch('/api/generate',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(payload)});
    const data=await res.json();
    document.getElementById('progressBar').style.width='100%';
    document.getElementById('progressText').textContent='Done!';
    showResults(data);
    toast(`${data.results.filter(r=>r.status==='ok').length} letters generated!`,'success');
  }catch(e){toast('Generation failed: '+e.message,'error');}
  finally{btn.disabled=false;btn.innerHTML='⚡ Generate All Offer Letters';setTimeout(()=>{document.getElementById('progressSection').style.display='none';},2000);}
}
function showResults(data){
  const card=document.getElementById('resultsCard');card.style.display='block';card.scrollIntoView({behavior:'smooth',block:'start'});
  const ok=data.results.filter(r=>r.status==='ok');
  if(ok.length){document.getElementById('downloadAllWrap').innerHTML=`<a href="/download-zip" class="btn btn-primary btn-full" style="margin-bottom:10px">📥 Download All ${ok.length} Letters (ZIP)</a>`;}
  let html='';
  data.results.forEach(r=>{
    const badge=r.status==='ok'?`<span class="pill pill-green">✓ Generated</span>`:`<span class="pill pill-red">✗ Error</span>`;
    const dl=r.status==='ok'?`<a href="/download/${r.intern_id}" class="btn btn-secondary btn-sm">⬇ PDF</a>`:`<span style="font-size:12px;color:var(--red)">${r.error||''}</span>`;
    html+=`<div class="result-item"><div><div class="result-name">${r.name}</div><div class="result-id">${r.intern_id}</div></div><div style="display:flex;align-items:center;gap:10px">${badge}${dl}</div></div>`;
  });
  document.getElementById('resultsList').innerHTML=html;
  if(data.email_log&&data.email_log.length){
    document.getElementById('emailReport').style.display='block';
    const sent=data.email_log.filter(e=>e.status==='sent').length,failed=data.email_log.filter(e=>e.status==='failed').length,skipped=data.email_log.filter(e=>e.status==='skipped').length;
    document.getElementById('emailStats').innerHTML=`<div class="card" style="text-align:center;margin-bottom:0"><div style="font-size:28px;font-weight:800;color:var(--green)">${sent}</div><div style="font-size:12px;color:var(--text3);margin-top:4px">Sent</div></div><div class="card" style="text-align:center;margin-bottom:0"><div style="font-size:28px;font-weight:800;color:var(--red)">${failed}</div><div style="font-size:12px;color:var(--text3);margin-top:4px">Failed</div></div><div class="card" style="text-align:center;margin-bottom:0"><div style="font-size:28px;font-weight:800;color:var(--yellow)">${skipped}</div><div style="font-size:12px;color:var(--text3);margin-top:4px">Skipped</div></div>`;
    let tbl=`<table><thead><tr><th>Name</th><th>Email</th><th>Status</th><th>Note</th></tr></thead><tbody>`;
    data.email_log.forEach(e=>{const cls=e.status==='sent'?'pill-green':e.status==='failed'?'pill-red':'pill-yellow';tbl+=`<tr><td class="name">${e.name}</td><td>${e.email}</td><td><span class="pill ${cls}">${e.status}</span></td><td style="font-size:12px;color:var(--red)">${e.error||''}</td></tr>`;});
    tbl+='</tbody></table>';document.getElementById('emailTable').innerHTML=tbl;
  }
}
</script>
</body>
</html>"""

@app.route("/")
def index(): return HTML

@app.route("/api/public-url")
def api_public_url():
    return jsonify({"url": PUBLIC_URL})

@app.route("/api/parse-excel", methods=["POST"])
def parse_excel():
    if "file" not in request.files: return jsonify({"error":"No file"}),400
    f=request.files["file"]
    try:
        df=pd.read_excel(f); df.columns=[c.strip() for c in df.columns]
        for col in ["Intern_id","Intern_ID","InternID"]:
            if col in df.columns: df=df.drop(columns=[col])
        missing=REQUIRED_COLS-set(df.columns)
        if missing: return jsonify({"error":f"Missing columns: {', '.join(sorted(missing))}"}),400
        interns=[{c:str(row[c]).strip() for c in df.columns}|{"_idx":i} for i,row in df.iterrows()]
        return jsonify({"success":True,"interns":interns,"count":len(interns)})
    except Exception as e: return jsonify({"error":str(e)}),500

@app.route("/api/upload-template", methods=["POST"])
def upload_template():
    global _template_bytes
    if "file" not in request.files: return jsonify({"error":"No file"}),400
    f=request.files["file"]
    if not f.filename.lower().endswith(".pdf"): return jsonify({"error":"Must be a PDF"}),400
    _template_bytes=f.read()
    return jsonify({"success":True})

@app.route("/api/test-email", methods=["POST"])
def test_email_route():
    data=request.json; smtp_cfg=data.get("smtp") or {}
    try:
        recipient=smtp_cfg.get("user","").strip()
        if not recipient: return jsonify({"error":"Enter your sender email address first."}),400
        send_email(cfg=smtp_cfg,recipient=recipient,intern_name="Test",pdf_bytes=b"%PDF-1.4 test",filename="test.pdf",subject="Reintenspark — SMTP Test",body="SMTP is working correctly!")
        return jsonify({"success":True})
    except Exception as e: return jsonify({"error":str(e)}),500

@app.route("/api/generate", methods=["POST"])
def generate_route():
    global _template_bytes
    if not _template_bytes: return jsonify({"error":"No template"}),400
    data=request.json
    interns=data.get("interns",[])
    id_prefix=data.get("id_prefix","RIS_")
    id_start=int(data.get("id_start",1))
    base_url=data.get("base_url","")
    do_email=data.get("send_email",False)
    smtp_cfg=data.get("smtp") or {}
    gen_dir=DATA_DIR/"generated"; gen_dir.mkdir(exist_ok=True)
    for p in gen_dir.glob("*.pdf"): p.unlink()
    results=[]; email_log=[]
    for i,intern in enumerate(interns):
        intern_id=f"{id_prefix}{id_start+i:06d}"
        intern["InternID"]=intern_id; name=intern.get("Name","")
        try:
            pdf_bytes=create_letter(intern,_template_bytes,base_url=base_url,include_qr=True)
            safe=name.replace(" ","_")
            filename=f"offer_letter_{intern_id}_{safe}.pdf"
            (gen_dir/filename).write_bytes(pdf_bytes)
            try: save_intern({**intern,"Status":"generated"})
            except: pass
            results.append({"intern_id":intern_id,"name":name,"filename":filename,"status":"ok"})
            if do_email and smtp_cfg:
                email=intern.get("Email","").strip()
                if email and email.lower() not in ("nan","none",""):
                    try:
                        send_email(cfg=smtp_cfg,recipient=email,intern_name=name,pdf_bytes=pdf_bytes,filename=filename,subject=smtp_cfg.get("subject","Offer Letter"),body=smtp_cfg.get("body","Dear {{Name}},"))
                        email_log.append({"name":name,"email":email,"status":"sent"})
                    except Exception as e:
                        email_log.append({"name":name,"email":email,"status":"failed","error":str(e)})
                else:
                    email_log.append({"name":name,"email":"—","status":"skipped"})
        except Exception as e:
            results.append({"intern_id":intern_id,"name":name,"status":"error","error":str(e)})
    return jsonify({"success":True,"results":results,"email_log":email_log if do_email else []})

@app.route("/download/<intern_id>")
def download_one(intern_id):
    gen_dir=DATA_DIR/"generated"
    matches=list(gen_dir.glob(f"offer_letter_{intern_id}_*.pdf"))
    if not matches: abort(404)
    return send_file(matches[0],mimetype="application/pdf",as_attachment=True,download_name=matches[0].name)

@app.route("/download-zip")
def download_zip():
    gen_dir=DATA_DIR/"generated"; pdfs=list(gen_dir.glob("*.pdf"))
    if not pdfs: abort(404)
    buf=io.BytesIO()
    with ZipFile(buf,"w") as zf:
        for p in pdfs: zf.write(p,arcname=p.name)
    buf.seek(0)
    return send_file(buf,mimetype="application/zip",as_attachment=True,download_name="reintenspark_offer_letters.zip")

@app.route("/verify/<intern_id>")
def verify_letter(intern_id):
    gen_dir=DATA_DIR/"generated"
    matches=list(gen_dir.glob(f"offer_letter_{intern_id}_*.pdf"))
    if not matches:
        return f"""<!DOCTYPE html><html><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1"><title>Not Found</title>
<style>body{{font-family:Arial,sans-serif;background:#f8fafc;display:flex;align-items:center;justify-content:center;min-height:100vh;margin:0}}
.box{{background:#fff;border-radius:16px;padding:48px 32px;text-align:center;box-shadow:0 4px 24px rgba(0,0,0,.08);max-width:400px;width:90%}}
.icon{{font-size:56px;margin-bottom:16px}}h2{{color:#ef4444;font-size:22px;margin-bottom:10px}}p{{color:#64748b;font-size:14px}}
.id{{font-family:monospace;background:#f1f5f9;padding:4px 10px;border-radius:6px;font-size:13px;color:#475569;margin-top:12px;display:inline-block}}</style></head>
<body><div class="box"><div class="icon">❌</div><h2>Document Not Found</h2><p>No verified offer letter found for this ID.</p><div class="id">{intern_id}</div></div></body></html>""",404
    fname=matches[0].name
    display_name=fname.replace(f"offer_letter_{intern_id}_","").replace(".pdf","").replace("_"," ")
    return f"""<!DOCTYPE html><html><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1"><title>Verified — Reintenspark</title>
<style>*{{margin:0;padding:0;box-sizing:border-box}}body{{font-family:'Segoe UI',sans-serif;background:linear-gradient(135deg,#f0fdf4,#e0f2fe);display:flex;align-items:center;justify-content:center;min-height:100vh}}
.box{{background:#fff;border-radius:20px;padding:48px 40px;text-align:center;box-shadow:0 8px 32px rgba(0,0,0,.10);max-width:440px;width:90%;border-top:5px solid #5CB85C}}
.icon{{font-size:64px;margin-bottom:20px}}.badge{{display:inline-flex;align-items:center;gap:8px;background:#f0fdf4;color:#166534;border:1.5px solid #bbf7d0;border-radius:30px;padding:6px 18px;font-size:13px;font-weight:700;margin-bottom:20px}}
h2{{color:#0f172a;font-size:24px;font-weight:800;margin-bottom:8px}}.name{{color:#5CB85C;font-size:20px;font-weight:700;margin-bottom:6px}}
.id{{font-family:monospace;background:#f1f5f9;padding:5px 14px;border-radius:8px;font-size:13px;color:#475569;display:inline-block;margin-bottom:24px}}
p{{color:#64748b;font-size:14px;line-height:1.6;margin-bottom:28px}}
.btn{{display:inline-flex;align-items:center;gap:8px;background:#5CB85C;color:#fff;padding:14px 32px;border-radius:10px;text-decoration:none;font-weight:700;font-size:15px}}
.btn:hover{{background:#4cae4c}}.footer{{margin-top:28px;font-size:11px;color:#94a3b8}}</style></head>
<body><div class="box"><div class="icon">✅</div><div class="badge">✓ VERIFIED DOCUMENT</div>
<h2>Offer Letter Verified</h2><div class="name">{display_name}</div><div class="id">{intern_id}</div>
<p>This offer letter has been issued by<br><strong>Reintenspark Technology Private Limited</strong><br>and is authentic.</p>
<a href="/view/{intern_id}" class="btn">📄 View Offer Letter</a>
<div class="footer">Reintenspark Technology Pvt. Ltd. · Document Verification System</div>
</div></body></html>"""

@app.route("/view/<intern_id>")
def view_letter(intern_id):
    gen_dir=DATA_DIR/"generated"
    matches=list(gen_dir.glob(f"offer_letter_{intern_id}_*.pdf"))
    if not matches: abort(404)
    doc=fitz.open(str(matches[0]))
    page=doc.load_page(0)
    page.add_redact_annot(fitz.Rect(QR_WIPE_X0,QR_WIPE_Y0,QR_WIPE_X1,QR_WIPE_Y1))
    page.apply_redactions(images=fitz.PDF_REDACT_IMAGE_NONE)
    buf=io.BytesIO()
    doc.save(buf,deflate=True,garbage=4,clean=True)
    doc.close(); buf.seek(0)
    return send_file(buf,mimetype="application/pdf",as_attachment=False,download_name=matches[0].name)

@app.route("/sample-excel")
def sample_excel():
    wb=Workbook(); ws=wb.active; ws.title="Interns"
    ws.append(["Name","Email","Domain","Mode","Start_date","Duration","End_Date"])
    ws.append(["Rahul Sharma","rahul@example.com","Web Development","Online","2025-03-01","3","2025-05-31"])
    ws.append(["Priya Patel","priya@example.com","Data Science","Offline","2025-03-01","6","2025-08-31"])
    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf,mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",as_attachment=True,download_name="sample_interns.xlsx")

# ══════════════════════════════════════════════════════════════════════════
import os

if __name__ == "__main__":
    init_excel()

    PORT = int(os.environ.get("PORT", 5000))

    print("\n🚀 Reintenspark Offer Letter System Running\n")

    app.run(host="0.0.0.0", port=PORT)